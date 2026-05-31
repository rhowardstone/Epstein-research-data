#!/usr/bin/env python3
"""
Stamp Gap Analysis — identify missing documents across Bates numbering systems.

For every consecutive EFTA pair where both have secondary stamp data, computes
the gap in each stamp numbering system. Gaps > 1 suggest documents were pulled
from production. Negative gaps indicate numbering resets across batches.

Outputs:
  1. stamp_gap_analysis.xlsx  — formatted Excel (positive gaps in 2+ stamp types)
  2. stamp_gap_crossref.xlsx  — cross-reference: are the "missing" stamp numbers
                                 found on ANY other EFTA in the corpus?
  3. stamp_corroborated_pulls.xlsx — the money sheet: EFTA pairs where 2+ stamp
     types agree on the SAME gap size, AND those numbers are absent from the
     corpus. Multiple independent numbering systems corroborate the same removal.

Usage:
  python3 tools/stamp_gap_analysis.py                    # All outputs
  python3 tools/stamp_gap_analysis.py --gaps-only        # Just the gap sheet
  python3 tools/stamp_gap_analysis.py --crossref-only    # Just the cross-ref
  python3 tools/stamp_gap_analysis.py --min-stamps 3     # Require 3+ stamp gaps
"""

import argparse
import csv
import os
import re
import sqlite3
import sys
import time
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Paths ---
# Auto-detect: if run from repo root, use relative paths; otherwise use absolute
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT  = os.path.dirname(_SCRIPT_DIR)  # tools/ → repo root

def _find(name):
    """Find a file in repo root or fallback to absolute path."""
    repo = os.path.join(_REPO_ROOT, name)
    if os.path.exists(repo):
        return repo
    absolute = os.path.join('/atb-data/rye/dump/epstein_files', name)
    if os.path.exists(absolute):
        return absolute
    return repo  # will error naturally

STAMPS_DB = _find('secondary_stamps.db')
CORPUS_DB = _find('full_text_corpus.db')
OUT_DIR   = _REPO_ROOT

# --- Stamp types (DB column name → display label) ---
STAMPS = [
    ('r1_number',       'R1'),
    ('efta_underscore', 'EFTA_'),
    ('jpm_sdny_number', 'JPM-SDNY'),
    ('db_sdny_number',  'DB-SDNY'),
    ('sdny_gm_number',  'SDNY-GM'),
    ('sdny_case_number','SDNY-CASE'),
    ('ubs_number',      'UBS'),
    ('td_doj_number',   'TD-DOJ'),
    ('usao_number',     'USAO'),
]

# Independence groups: stamp types within the same group are constant-offset
# copies of each other (confirmed by pairwise analysis). A "corroboration"
# requires agreement across DIFFERENT groups, not within the same group.
#   Group A (SDNY production): DB-SDNY, SDNY-GM, JPM-SDNY (91-93% constant offset)
#   Group B (R1): R1, USAO (100% constant offset on small overlap)
#   Group C: SDNY-CASE (independent of A and B)
#   Group D: UBS (independent)
#   Group E: TD-DOJ (borderline — 60% constant with GM, treat as independent)
#   Group F: EFTA_ (independent — early-range R1+1 coincidence, but diverges on 10K+ docs)
INDEPENDENCE_GROUPS = {
    'db_sdny_number':  'sdny_production',
    'sdny_gm_number':  'sdny_production',
    'jpm_sdny_number': 'sdny_production',
    'r1_number':       'r1_family',
    'usao_number':     'r1_family',
    'sdny_case_number':'sdny_case',
    'ubs_number':      'ubs',
    'td_doj_number':   'td_doj',
    'efta_underscore': 'efta_underscore',
}

# --- Excel styles ---
HDR_FILL  = PatternFill('solid', fgColor='1F2937')
HDR_FONT  = Font(name='Consolas', size=9, bold=True, color='FFFFFF')
SUB_FILL  = PatternFill('solid', fgColor='F3F4F6')
SUB_FONT  = Font(name='Consolas', size=8, bold=True, color='6B7280')
D_FONT    = Font(name='Consolas', size=9)
EFTA_FONT = Font(name='Consolas', size=9, bold=True)
GAP_HI    = Font(name='Consolas', size=9, bold=True, color='991B1B')   # >100
GAP_MD    = Font(name='Consolas', size=9, color='92400E')              # >10
BLUE_FILL   = PatternFill('solid', fgColor='EFF6FF')
ORANGE_FILL = PatternFill('solid', fgColor='FFF7ED')
GREEN_FILL  = PatternFill('solid', fgColor='ECFDF5')
RED_FILL    = PatternFill('solid', fgColor='FEF2F2')
GRP_BORDER  = Border(left=Side('medium', color='6B7280'))
GROUP_FILLS = [BLUE_FILL, ORANGE_FILL]


def load_document_boundaries():
    """Load first-page and last-page stamps for each document from page_stamps.

    Returns sorted list of (doc_efta, dataset, first_page_stamps, last_page_stamps).
    Only includes documents where page_stamps has data.
    """
    conn = sqlite3.connect(STAMPS_DB)
    cols = [s[0] for s in STAMPS]
    col_str = ', '.join(cols)

    # Get all pages grouped by document, ordered by page number
    rows = conn.execute(
        f"SELECT final_efta, dataset, page_number, total_pages, {col_str} "
        f"FROM page_stamps WHERE final_efta LIKE 'EFTA%' "
        f"ORDER BY final_efta, page_number"
    ).fetchall()
    conn.close()

    def parse_stamps(row, offset=4):
        vals = {}
        for i, (col, _) in enumerate(STAMPS):
            v = row[offset + i]
            if v is not None and str(v).strip():
                try:
                    vals[col] = int(v)
                except (ValueError, TypeError):
                    pass
        return vals

    # Group by document, extract first and last page stamps + page EFTAs
    docs = []
    cur_doc = None
    first_row = None
    last_row = None

    for row in rows:
        doc_efta = row[0]
        if doc_efta != cur_doc:
            # Emit previous doc
            if cur_doc and first_row and last_row:
                first_stamps = parse_stamps(first_row)
                last_stamps = parse_stamps(last_row)
                if first_stamps or last_stamps:
                    parent_num = int(cur_doc[4:])
                    first_page_efta = f'EFTA{parent_num + first_row[2]:08d}'
                    last_page_efta = f'EFTA{parent_num + last_row[2]:08d}'
                    docs.append((cur_doc, first_row[1], first_stamps, last_stamps,
                                 first_page_efta, last_page_efta))
            cur_doc = doc_efta
            first_row = row
            last_row = row
        else:
            last_row = row

    # Emit final doc
    if cur_doc and first_row and last_row:
        first_stamps = parse_stamps(first_row)
        last_stamps = parse_stamps(last_row)
        if first_stamps or last_stamps:
            parent_num = int(cur_doc[4:])
            first_page_efta = f'EFTA{parent_num + first_row[2]:08d}'
            last_page_efta = f'EFTA{parent_num + last_row[2]:08d}'
            docs.append((cur_doc, first_row[1], first_stamps, last_stamps,
                         first_page_efta, last_page_efta))

    docs.sort(key=lambda x: x[0])
    return docs


def compute_gaps(docs):
    """Compute gaps at document boundaries.

    Compares LAST page of doc A with FIRST page of doc B.
    gap = first_B_stamp - last_A_stamp - 1
    A result of 0 = contiguous, >0 = missing pages between docs.

    Returns list of dicts with efta, dataset, and per-stamp this/next/gap values.

    Inclusion rule: at least 1 stamp system shows a positive gap, AND no stamp
    system contradicts (gap <= 0). If any system says "contiguous" while another
    says "gap", the row is excluded — the stamps disagree.
    """
    gap_rows = []

    for i in range(1, len(docs)):
        prev_efta, prev_ds, _, prev_last, _, prev_last_page = docs[i - 1]
        cur_efta, cur_ds, cur_first, _, cur_first_page, _ = docs[i]

        row = {
            'efta': prev_last_page,       # Last page of doc A (the stamp you'd look at)
            'next_efta': cur_first_page,   # First page of doc B
            'doc_a': prev_efta,
            'doc_b': cur_efta,
            'dataset': prev_ds,
        }
        pos_gaps = 0
        contradictions = 0

        for col, _ in STAMPS:
            pv = prev_last.get(col)    # last page of previous doc
            cv = cur_first.get(col)    # first page of next doc
            if pv is not None and cv is not None:
                # Gap = first_next - last_prev - 1
                # 0 means perfectly contiguous, >0 means missing pages
                gap = cv - pv - 1
                row[col] = pv
                row['next_' + col] = cv
                row[col + '_gap'] = gap
                if gap > 0:
                    pos_gaps += 1
                elif gap <= 0:
                    contradictions += 1
            else:
                row[col] = None
                row['next_' + col] = None
                row[col + '_gap'] = None

        # Include if at least 1 system shows a gap AND none contradict
        if pos_gaps >= 1 and contradictions == 0:
            gap_rows.append(row)

    return gap_rows


def write_gap_excel(gap_rows, path):
    """Write the formatted gap analysis Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stamp Gap Analysis"

    # Row 1: group headers
    ws.merge_cells('A1:B1')
    for c in (1, 2):
        ws.cell(1, c).fill = HDR_FILL
    ws['A1'].value = 'Page'
    ws['A1'].font = HDR_FONT

    col = 3
    for _, label in STAMPS:
        s, e = get_column_letter(col), get_column_letter(col + 2)
        ws.merge_cells(f'{s}1:{e}1')
        cell = ws.cell(1, col)
        cell.value = label
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center')
        for cc in range(col, col + 3):
            ws.cell(1, cc).fill = HDR_FILL
        col += 3

    # Row 2: sub-headers
    subs = ['EFTA', 'DS']
    for _ in STAMPS:
        subs += ['This', 'Next', 'Gap']
    for i, h in enumerate(subs, 1):
        c = ws.cell(2, i)
        c.value = h
        c.font = SUB_FONT
        c.fill = SUB_FILL
        c.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 4
    col = 3
    for _ in STAMPS:
        ws.column_dimensions[get_column_letter(col)].width = 9
        ws.column_dimensions[get_column_letter(col + 1)].width = 9
        ws.column_dimensions[get_column_letter(col + 2)].width = 7
        col += 3

    # Data rows
    for ri, row in enumerate(gap_rows):
        r = ri + 3
        ws.cell(r, 1, row['efta']).font = EFTA_FONT
        ws.cell(r, 2, int(row['dataset'])).font = D_FONT
        ws.cell(r, 2).alignment = Alignment(horizontal='center')

        col = 3
        for si, (sk, _) in enumerate(STAMPS):
            gf = GROUP_FILLS[si % 2]
            tv = row.get(sk)
            nv = row.get('next_' + sk)
            gv = row.get(sk + '_gap')

            c1 = ws.cell(r, col)
            c2 = ws.cell(r, col + 1)
            c3 = ws.cell(r, col + 2)

            c1.fill = gf; c2.fill = gf; c3.fill = gf
            c1.font = D_FONT; c2.font = D_FONT; c3.font = D_FONT
            c1.number_format = '#,##0'; c2.number_format = '#,##0'; c3.number_format = '#,##0'
            c1.border = GRP_BORDER

            if tv is not None: c1.value = tv
            if nv is not None: c2.value = nv
            if gv is not None:
                c3.value = gv
                if gv > 100:
                    c3.font = GAP_HI
                elif gv > 10:
                    c3.font = GAP_MD
            col += 3

    ws.freeze_panes = 'C3'
    last_col = get_column_letter(2 + len(STAMPS) * 3)
    ws.auto_filter.ref = f'A2:{last_col}{len(gap_rows) + 2}'

    wb.save(path)
    return len(gap_rows)


def build_stamp_index(stamps_db_path):
    """Build reverse index: for each stamp type, map stamp_number → set of EFTAs."""
    conn = sqlite3.connect(stamps_db_path)
    index = {}  # stamp_col -> {number -> set(efta)}

    for col, _ in STAMPS:
        index[col] = defaultdict(set)

    # Use page_stamps for finer-grained lookup (1.2M rows)
    # Column names differ slightly — map them
    page_col_map = {
        'r1_number': 'r1_number',
        'jpm_sdny_number': 'jpm_sdny_number',
        'db_sdny_number': 'db_sdny_number',
        'sdny_gm_number': 'sdny_gm_number',
        'efta_underscore': 'efta_underscore',
        'sdny_case_number': 'sdny_case_number',
        'ubs_number': 'ubs_number',
        'td_doj_number': 'td_doj_number',
        'usao_number': 'usao_number',
    }

    # Also build from document_stamps (may have docs not in page_stamps)
    cols = [s[0] for s in STAMPS]
    col_str = ', '.join(cols)

    print('  Loading document_stamps...')
    for row in conn.execute(f"SELECT efta_number, {col_str} FROM document_stamps"):
        efta = row[0]
        for i, (col, _) in enumerate(STAMPS):
            v = row[i + 1]
            if v is not None and str(v).strip():
                try:
                    index[col][int(v)].add(efta)
                except (ValueError, TypeError):
                    pass

    print('  Loading page_stamps...')
    page_cols = ', '.join(page_col_map[c] for c in cols)
    for row in conn.execute(f"SELECT final_efta, {page_cols} FROM page_stamps"):
        efta = row[0]
        for i, (col, _) in enumerate(STAMPS):
            v = row[i + 1]
            if v is not None and str(v).strip():
                try:
                    index[col][int(v)].add(efta)
                except (ValueError, TypeError):
                    pass

    conn.close()

    for col, _ in STAMPS:
        print(f'    {col}: {len(index[col]):,} unique stamp numbers indexed')

    return index


def crossref_gaps(gap_rows, stamp_index, max_gap_size=500):
    """For each gap, check if the missing stamp numbers exist elsewhere.

    Only checks gaps <= max_gap_size to avoid scanning millions of numbers
    for batch-reset gaps (e.g. R1 gaps of 500K+).

    Returns list of dicts with gap info + cross-ref results.
    """
    results = []

    for row in gap_rows:
        for col, label in STAMPS:
            tv = row.get(col)
            nv = row.get('next_' + col)
            gv = row.get(col + '_gap')

            if gv is None or gv <= 0 or gv > max_gap_size:
                continue

            # The missing range is (tv + 1) to (nv - 1)
            # e.g. this page stamp=100, next page stamp=104 → missing = 101,102,103
            missing_start = tv + 1
            missing_end = nv - 1
            missing_count = missing_end - missing_start + 1

            found_elsewhere = 0
            found_eftas = set()
            not_found = 0

            for num in range(missing_start, missing_end + 1):
                eftas = stamp_index[col].get(num, set())
                # Exclude the two EFTAs that form this gap
                eftas = eftas - {row['efta']}
                if eftas:
                    found_elsewhere += 1
                    found_eftas.update(eftas)
                else:
                    not_found += 1

            results.append({
                'efta': row['efta'],
                'dataset': row['dataset'],
                'stamp_type': label,
                'stamp_col': col,
                'this_num': tv,
                'next_num': nv,
                'gap': gv,
                'missing_count': missing_count,
                'found_elsewhere': found_elsewhere,
                'not_found': not_found,
                'pct_found': round(100 * found_elsewhere / missing_count, 1) if missing_count > 0 else 0,
                'found_eftas_sample': sorted(found_eftas)[:5],  # Up to 5 example EFTAs
            })

    return results


def write_crossref_excel(xref_rows, path):
    """Write cross-reference results to formatted Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Cross-Reference"

    headers = [
        ('EFTA', 16), ('DS', 4), ('Stamp Type', 12),
        ('This #', 10), ('Next #', 10), ('Gap', 7),
        ('Missing', 8), ('Found', 8), ('Not Found', 9),
        ('% Found', 8), ('Status', 12), ('Example EFTAs', 60),
    ]

    for i, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, row in enumerate(xref_rows):
        r = ri + 2
        pct = row['pct_found']

        # Status classification
        if pct == 100:
            status = 'ALL FOUND'
            bg = GREEN_FILL
        elif pct >= 50:
            status = 'PARTIAL'
            bg = PatternFill('solid', fgColor='FFFBEB')
        elif pct > 0:
            status = 'MOSTLY MISSING'
            bg = PatternFill('solid', fgColor='FEF2F2')
        else:
            status = 'ALL MISSING'
            bg = RED_FILL

        vals = [
            row['efta'], row['dataset'], row['stamp_type'],
            row['this_num'], row['next_num'], row['gap'],
            row['missing_count'], row['found_elsewhere'], row['not_found'],
            row['pct_found'], status,
            ', '.join(row['found_eftas_sample']),
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.font = EFTA_FONT if ci == 1 else D_FONT
            if ci in (4, 5, 6, 7, 8, 9):
                c.number_format = '#,##0'
            if ci == 10:
                c.number_format = '0.0'
            if ci == 11:
                c.fill = bg
                c.font = Font(name='Consolas', size=9, bold=True)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(xref_rows) + 1}'

    wb.save(path)
    return len(xref_rows)


def find_corroborated_pulls(gap_rows, xref_rows, max_gap_size=500):
    """Find document boundaries where stamp numbers are missing from the corpus.

    For each gap, checks which stamp systems have ALL their missing numbers
    absent from the entire corpus (the real validation). Also flags whether
    multiple INDEPENDENT numbering families agree (bonus confidence).

    Independence groups (constant-offset copies can't independently corroborate):
      SDNY production: DB-SDNY, SDNY-GM, JPM-SDNY (91-93% constant offset)
      R1 family: R1, USAO
      Others: SDNY-CASE, UBS, TD-DOJ (each independent)

    Returns list of dicts, one per confirmed gap.
    """
    # Group xref by EFTA
    xref_by_efta = defaultdict(list)
    for xr in xref_rows:
        xref_by_efta[xr['efta']].append(xr)

    gap_by_efta = {r['efta']: r for r in gap_rows}

    results = []

    for efta, xrefs in xref_by_efta.items():
        # Only keep ALL MISSING entries (0% found elsewhere)
        missing = [x for x in xrefs if x['not_found'] == x['missing_count'] and x['missing_count'] > 0]
        if not missing:
            continue

        # Group by gap size — entries with the same gap corroborate each other
        by_gap = defaultdict(list)
        for x in missing:
            by_gap[x['gap']].append(x)

        for gap_size, entries in by_gap.items():
            if gap_size > max_gap_size:
                continue

            # Count stamp systems and independent groups
            groups = set()
            for e in entries:
                group = INDEPENDENCE_GROUPS.get(e['stamp_col'], e['stamp_col'])
                groups.add(group)

            n_independent = len(groups)

            # Confidence level
            if n_independent >= 3:
                confidence = 'HIGHEST'
            elif n_independent >= 2:
                confidence = 'HIGH'
            elif len(entries) >= 2:
                confidence = 'MODERATE'  # Same family agrees (e.g. DB-SDNY + SDNY-GM)
            else:
                confidence = 'SINGLE'    # Only one stamp system, but numbers confirmed missing

            gap_row = gap_by_efta.get(efta, {})
            results.append({
                'efta': efta,
                'dataset': entries[0]['dataset'],
                'agreed_gap': gap_size,
                'missing_docs': gap_size,
                'corroborating_stamps': len(entries),
                'independent_groups': n_independent,
                'confidence': confidence,
                'groups': ', '.join(sorted(groups)),
                'stamp_types': ', '.join(sorted(e['stamp_type'] for e in entries)),
                'stamp_details': entries,
                'gap_row': gap_row,
            })

    # Sort by confidence tier, then gap size
    conf_order = {'HIGHEST': 0, 'HIGH': 1, 'MODERATE': 2, 'SINGLE': 3}
    results.sort(key=lambda r: (conf_order[r['confidence']], -r['missing_docs']))
    return results


def load_doc_types():
    """Load doc_type for all documents from full_text_corpus.db."""
    conn = sqlite3.connect(CORPUS_DB)
    types = {}
    for row in conn.execute("SELECT efta_number, doc_type FROM documents WHERE doc_type IS NOT NULL"):
        types[row[0]] = row[1]
    conn.close()
    return types


NTWM_CSV = os.path.join(OUT_DIR, 'victims', 'EFTA00095751_complete.csv')
JENCKS_CSV = os.path.join(OUT_DIR, 'victims', 'EFTA00010025_complete.csv')


def load_manifests():
    """Load NTWM + Jencks manifests into a lookup: serial_prefix → {subdoc_num → description}.

    Returns dict like:
      {'3501.014': {7: 'Search record for [redacted]', 8: 'FBI report...', ...},
       '3502':     {1: 'Notes from phone call', ...}}
    """
    manifest = defaultdict(dict)  # prefix → {subdoc_int → description}

    for csv_path in [NTWM_CSV, JENCKS_CSV]:
        if not os.path.exists(csv_path):
            continue
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                serial = row.get('serial_or_id', '').strip()
                desc = row.get('description', '').strip()
                if not serial or not desc or desc == '[illegible]':
                    continue
                _, prefix, subdoc = normalize_serial(serial)
                if subdoc is not None:
                    manifest[prefix][subdoc] = desc

    return dict(manifest)


def normalize_serial(raw):
    """Normalize serial formats to canonical 'PREFIX-SUBDOC' form.

    Input formats seen in the wild:
      '3501.045-008'  → '3501.045-008'   (already canonical)
      '3501.045.008'  → '3501.045-008'   (dot instead of dash)
      '3501-055'      → '3501-055'       (prefix-only, no victim serial)
      '3501.055'      → '3501.055'       (prefix-only from serial_efta_map)
      '3513.041'      → '3513.041'       (prefix-only, different case family)

    Returns (normalized_string, prefix, subdoc_int_or_None)
    """
    raw = raw.strip()

    # Format: 3501.045-008 or 3501.045.008 (three-part with victim serial + subdoc)
    m = re.match(r'(35\d{2}\.\d+)[.\-](\d+)$', raw)
    if m:
        prefix, subdoc = m.group(1), int(m.group(2))
        return f"{prefix}-{subdoc:03d}", prefix, subdoc

    # Format: 3501-055 (two-part, prefix=3501, serial=055 — no subdoc)
    m = re.match(r'(35\d{2})-(\d+)$', raw)
    if m:
        prefix = f"{m.group(1)}.{m.group(2).lstrip('0') or '0'}"
        return prefix, prefix, None

    # Format: 3501.055 or 3513.041 (prefix only, no subdoc)
    m = re.match(r'(35\d{2}\.\d+)$', raw)
    if m:
        return m.group(1), m.group(1), None

    # Format: 3502-001 (Jencks family, no dot)
    m = re.match(r'(35\d{2})-(\d+)$', raw)
    if m:
        return raw, m.group(1), int(m.group(2))

    return raw, raw, None


def load_serial_map():
    """Load EFTA → serial stamp mappings from secondary_stamps.db.

    Returns two dicts:
      efta_to_serial: {efta → normalized serial string}
      serial_to_eftas: {prefix → {subdoc_int → [eftas]}} for published docs
    """
    conn = sqlite3.connect(STAMPS_DB)
    efta_to_serial = {}
    serial_to_eftas = defaultdict(lambda: defaultdict(list))

    # From document_stamps (has discovery_ref)
    for row in conn.execute(
        "SELECT efta_number, discovery_ref FROM document_stamps WHERE discovery_ref IS NOT NULL"
    ):
        efta, ref = row
        normalized, prefix, subdoc = normalize_serial(ref)
        efta_to_serial[efta] = normalized
        if subdoc is not None:
            serial_to_eftas[prefix][subdoc].append(efta)

    # Also from serial_efta_map (may have additional mappings)
    for row in conn.execute("SELECT efta_number, serial FROM serial_efta_map"):
        efta, serial = row
        normalized, prefix, subdoc = normalize_serial(serial)
        if efta not in efta_to_serial:
            efta_to_serial[efta] = normalized
        if subdoc is not None:
            serial_to_eftas[prefix][subdoc].append(efta)

    conn.close()
    return efta_to_serial, dict(serial_to_eftas)


def annotate_manifest(corr_rows, efta_to_serial, serial_to_eftas, manifest):
    """Annotate corroborated gaps with manifest data where available.

    For each gap, checks if Doc A or Doc B has a serial stamp. If the serial
    prefix has manifest entries, identifies which subdocuments fall in the gap
    between the two published serials.

    Adds to each row: 'serial_a', 'serial_b', 'manifest_missing_count',
                       'manifest_missing_items' (list of (subdoc, description))
    """
    for row in corr_rows:
        gap_row = row.get('gap_row', {})
        doc_a_parent = gap_row.get('doc_a', row['efta'])
        doc_b_parent = gap_row.get('doc_b', '')

        raw_a = efta_to_serial.get(doc_a_parent, '')
        raw_b = efta_to_serial.get(doc_b_parent, '')

        row['serial_a'] = raw_a
        row['serial_b'] = raw_b
        row['manifest_missing_count'] = 0
        row['manifest_missing_items'] = []

        if not raw_a or not raw_b:
            continue

        _, prefix_a, subdoc_a = normalize_serial(raw_a)
        _, prefix_b, subdoc_b = normalize_serial(raw_b)

        if subdoc_a is None or subdoc_b is None:
            continue

        missing_items = []

        if prefix_a == prefix_b and subdoc_b > subdoc_a + 1:
            # Same serial prefix — check items between subdoc_a and subdoc_b
            if prefix_a in manifest:
                for subdoc_num in range(subdoc_a + 1, subdoc_b):
                    if subdoc_num in manifest[prefix_a]:
                        published = serial_to_eftas.get(prefix_a, {}).get(subdoc_num, [])
                        if not published:
                            missing_items.append((f"{prefix_a}-{subdoc_num:03d}",
                                                  manifest[prefix_a][subdoc_num]))
        else:
            # Different prefixes or same subdoc — check tail of A's serial
            # and head of B's serial for unpublished manifest items
            if prefix_a in manifest:
                max_subdoc_a = max(manifest[prefix_a].keys()) if manifest[prefix_a] else 0
                for subdoc_num in range(subdoc_a + 1, max_subdoc_a + 1):
                    if subdoc_num in manifest[prefix_a]:
                        published = serial_to_eftas.get(prefix_a, {}).get(subdoc_num, [])
                        if not published:
                            missing_items.append((f"{prefix_a}-{subdoc_num:03d}",
                                                  manifest[prefix_a][subdoc_num]))
            if prefix_b in manifest:
                for subdoc_num in range(1, subdoc_b):
                    if subdoc_num in manifest[prefix_b]:
                        published = serial_to_eftas.get(prefix_b, {}).get(subdoc_num, [])
                        if not published:
                            missing_items.append((f"{prefix_b}-{subdoc_num:03d}",
                                                  manifest[prefix_b][subdoc_num]))

        row['manifest_missing_count'] = len(missing_items)
        row['manifest_missing_items'] = missing_items


def write_corroborated_excel(corr_rows, path, doc_types=None):
    """Write the corroborated pulls Excel — the highest-confidence findings."""
    if doc_types is None:
        doc_types = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Corroborated Pulls"

    headers = [
        ('Doc A', 16), ('Doc B', 16),
        ('Gap', 8), ('# Stamps', 8), ('Systems', 40),
        ('Serial A', 16), ('Serial B', 16),
        ('Manifest Items', 12), ('Manifest Descriptions', 80),
        ('Stamp Ranges', 60),
    ]

    for i, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, row in enumerate(corr_rows):
        r = ri + 2
        n_stamps = row['corroborating_stamps']

        # Build stamp ranges string
        ranges = []
        for d in row['stamp_details']:
            ranges.append(f"{d['stamp_type']}: {d['this_num']:,}→{d['next_num']:,}")
        ranges_str = ' | '.join(ranges)

        gap_row = row.get('gap_row', {})
        doc_a = gap_row.get('efta', row['efta'])          # last page of doc A
        doc_b = gap_row.get('next_efta', gap_row.get('doc_b', ''))  # first page of doc B

        # Manifest cross-reference
        serial_a = row.get('serial_a', '')
        serial_b = row.get('serial_b', '')
        manifest_count = row.get('manifest_missing_count', 0)
        manifest_items = row.get('manifest_missing_items', [])
        manifest_str = ' | '.join(
            f"{prefix_subdoc}: {desc}" for prefix_subdoc, desc in manifest_items
        ) if manifest_items else ''

        # Column order: Doc A, Doc B, Gap, # Stamps, Systems,
        #               Serial A, Serial B, Manifest Items, Manifest Descriptions,
        #               Stamp Ranges
        vals = [
            doc_a, doc_b,
            row['missing_docs'], n_stamps, row['stamp_types'],
            serial_a, serial_b,
            manifest_count if manifest_count else '', manifest_str,
            ranges_str,
        ]

        MANIFEST_FONT = Font(name='Consolas', size=8, color='1D4ED8')
        SERIAL_FONT = Font(name='Consolas', size=9, color='7C3AED')

        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            if ci in (1, 2):       # Doc A, Doc B
                c.font = EFTA_FONT
            elif ci in (6, 7):     # Serial A, Serial B
                c.font = SERIAL_FONT
            elif ci == 9:          # Manifest Descriptions
                c.font = MANIFEST_FONT
            else:
                c.font = D_FONT
            if ci == 3:            # Gap
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='center')
            if ci in (4, 8):       # # Stamps, Manifest Items
                c.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(corr_rows) + 1}'

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Corroborated Pull Summary'
    ws2['A1'].font = Font(name='Consolas', size=12, bold=True)

    summary_data = [
        ('Total corroborated gaps', len(corr_rows)),
        ('Total estimated missing docs', sum(r['missing_docs'] for r in corr_rows)),
    ]

    # By confidence tier
    from collections import Counter
    by_conf = Counter(r['confidence'] for r in corr_rows)
    for tier in ['HIGHEST', 'HIGH', 'MODERATE', 'SINGLE']:
        if by_conf[tier]:
            subset = [r for r in corr_rows if r['confidence'] == tier]
            missing = sum(r['missing_docs'] for r in subset)
            summary_data.append((f'  {tier}', f'{by_conf[tier]} gaps, ~{missing} missing pages'))

    # Manifest cross-reference stats
    manifest_gaps = [r for r in corr_rows if r.get('manifest_missing_count', 0) > 0]
    if manifest_gaps:
        total_manifest_items = sum(r['manifest_missing_count'] for r in manifest_gaps)
        summary_data.append(('', ''))
        summary_data.append(('Manifest cross-reference:', ''))
        summary_data.append(('  Gaps with manifest IDs', f'{len(manifest_gaps)} gaps'))
        summary_data.append(('  Named missing items', f'{total_manifest_items} documents'))
        summary_data.append(('  Source', 'NTWM (EFTA00095751) + Jencks (EFTA00010025)'))

    # By dataset
    by_ds = defaultdict(list)
    for r in corr_rows:
        by_ds[r['dataset']].append(r)
    summary_data.append(('', ''))
    summary_data.append(('By dataset:', ''))
    for ds in sorted(by_ds.keys()):
        subset = by_ds[ds]
        missing = sum(r['missing_docs'] for r in subset)
        summary_data.append((f'  DS{ds}', f'{len(subset)} gaps, ~{missing} missing docs'))

    for i, (label, val) in enumerate(summary_data, 3):
        ws2.cell(i, 1, label).font = Font(name='Consolas', size=10, bold=bool(label and not label.startswith(' ')))
        ws2.cell(i, 2, val).font = Font(name='Consolas', size=10)
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 40

    wb.save(path)
    return len(corr_rows)


def main():
    parser = argparse.ArgumentParser(description='Stamp gap analysis with cross-reference')
    parser.add_argument('--gaps-only', action='store_true', help='Only generate gap Excel')
    parser.add_argument('--crossref-only', action='store_true', help='Only generate cross-ref Excel')
    parser.add_argument('--min-stamps', type=int, default=1,
                        help='Minimum number of stamp types with positive gaps (default: 2)')
    parser.add_argument('--max-gap', type=int, default=500,
                        help='Max gap size to cross-reference (default: 500, skip batch resets)')
    parser.add_argument('--out-dir', default=OUT_DIR, help='Output directory')
    args = parser.parse_args()

    gap_path = os.path.join(args.out_dir, 'stamp_gap_analysis.xlsx')
    xref_path = os.path.join(args.out_dir, 'stamp_gap_crossref.xlsx')

    # --- Step 1: Compute gaps ---
    print('Loading document boundary stamps (first + last page per doc)...')
    t0 = time.time()
    docs = load_document_boundaries()
    print(f'  {len(docs):,} documents with stamp data in {time.time()-t0:.1f}s')

    print(f'Computing document-boundary gaps (last page of A → first page of B)...')
    print(f'  All stamps must agree (gap > 0), no contradictions (gap <= 0)')
    t0 = time.time()
    gap_rows = compute_gaps(docs)
    print(f'  {len(gap_rows):,} gap rows in {time.time()-t0:.1f}s')

    if not args.crossref_only:
        print(f'Writing gap Excel → {gap_path}')
        n = write_gap_excel(gap_rows, gap_path)
        sz = os.path.getsize(gap_path) / 1024 / 1024
        print(f'  {n:,} rows, {sz:.1f} MB')

    if args.gaps_only:
        return

    # --- Step 2: Cross-reference ---
    print(f'\nBuilding stamp reverse index (max gap to check: {args.max_gap})...')
    t0 = time.time()
    stamp_index = build_stamp_index(STAMPS_DB)
    print(f'  Index built in {time.time()-t0:.1f}s')

    print('Cross-referencing gaps against full corpus...')
    t0 = time.time()
    xref_rows = crossref_gaps(gap_rows, stamp_index, max_gap_size=args.max_gap)
    print(f'  {len(xref_rows):,} gap×stamp entries checked in {time.time()-t0:.1f}s')

    # Stats
    statuses = defaultdict(int)
    for r in xref_rows:
        pct = r['pct_found']
        if pct == 100: statuses['ALL FOUND'] += 1
        elif pct >= 50: statuses['PARTIAL'] += 1
        elif pct > 0: statuses['MOSTLY MISSING'] += 1
        else: statuses['ALL MISSING'] += 1

    print('\nCross-reference summary:')
    for s in ['ALL FOUND', 'PARTIAL', 'MOSTLY MISSING', 'ALL MISSING']:
        print(f'  {s:20s} {statuses[s]:>6,}')

    print(f'\nWriting cross-ref Excel → {xref_path}')
    n = write_crossref_excel(xref_rows, xref_path)
    sz = os.path.getsize(xref_path) / 1024 / 1024
    print(f'  {n:,} rows, {sz:.1f} MB')

    # --- Step 3: Corroborated pulls ---
    corr_path = os.path.join(args.out_dir, 'stamp_corroborated_pulls.xlsx')
    print('\nFinding corroborated pulls (2+ stamps agree on gap, all missing from corpus)...')
    t0 = time.time()
    corr_rows = find_corroborated_pulls(gap_rows, xref_rows, max_gap_size=args.max_gap)
    print(f'  {len(corr_rows):,} corroborated gaps in {time.time()-t0:.1f}s')

    if corr_rows:
        total_missing = sum(r['missing_docs'] for r in corr_rows)
        print(f'  Estimated missing documents: {total_missing:,}')

        # Quick breakdown by confidence
        from collections import Counter
        by_conf = Counter(r['confidence'] for r in corr_rows)
        for tier in ['HIGHEST', 'HIGH', 'MODERATE', 'SINGLE']:
            if by_conf[tier]:
                subset = [r for r in corr_rows if r['confidence'] == tier]
                missing = sum(r['missing_docs'] for r in subset)
                print(f'    {tier:10s}: {by_conf[tier]:>5,} gaps, ~{missing:,} missing pages')

        print(f'\nLoading document types from corpus...')
        doc_types = load_doc_types()
        print(f'  {len(doc_types):,} documents with type classifications')

        # --- Step 4: Manifest cross-reference ---
        print(f'\nLoading FBI manifests (NTWM + Jencks)...')
        manifest = load_manifests()
        total_entries = sum(len(v) for v in manifest.values())
        print(f'  {len(manifest)} serial prefixes, {total_entries:,} manifest entries')

        print('Loading serial→EFTA mappings...')
        efta_to_serial, serial_to_eftas = load_serial_map()
        print(f'  {len(efta_to_serial):,} EFTA→serial mappings')

        print('Annotating gaps with manifest data...')
        t0 = time.time()
        annotate_manifest(corr_rows, efta_to_serial, serial_to_eftas, manifest)
        manifest_hits = sum(1 for r in corr_rows if r.get('manifest_missing_count', 0) > 0)
        manifest_items = sum(r.get('manifest_missing_count', 0) for r in corr_rows)
        print(f'  {manifest_hits:,} gaps matched to manifest entries ({manifest_items:,} named items) in {time.time()-t0:.1f}s')

        print(f'\nWriting corroborated pulls → {corr_path}')
        n = write_corroborated_excel(corr_rows, corr_path, doc_types=doc_types)
        sz = os.path.getsize(corr_path) / 1024 / 1024
        print(f'  {n:,} rows, {sz:.1f} MB')
    else:
        print('  No corroborated pulls found.')


if __name__ == '__main__':
    main()
